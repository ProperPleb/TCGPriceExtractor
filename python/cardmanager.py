import json
import util
import config
import re

from rest import Rest
from card import Card
from openpyxl import load_workbook, Workbook


class CardManager:
    CARD_CONDITION_MAP = {"NM": "Near Mint",
                          "LP": "Lightly Played",
                          "MP": "Moderately Played",
                          "HP": "Heavily Played"}

    CARD_EDITION_MAP = {"F": "1st Edition",
                        "U": "Unlimited",
                        "L": "Limited"}

    CARD_RARITY_MAP = {"C": "Common / Short Print",
                       "R": "Rare",
                       "SP": "Super Rare",
                       "UR": "Ultra Rare",
                       "SEP": "Secret Rare",
                       "UTR": "Ultimate Rare",
                       "CR": "Collector's Rare",
                       "QR": "Quarter Century Secret Rare",
                       "STR": "Starlight Rare"}

    def __init__(self, file_path: str = None, file_name: str = None, rest: Rest = None, deck: tuple = None):
        self.file_path = file_path if file_path is not None else config.DEFAULT_PATH
        self.file_name = file_name if file_name is not None else 'catalog.xlsx'
        self.rest = rest if rest is not None else Rest()
        self.deck = deck

    def populate_deck(self):
        wb = load_workbook(self.file_path + self.file_name)
        dict_list = util.excel_mapper(wb)
        card_list = []
        for entry in dict_list:
            self.sanitize(entry)
            card_list.append(Card(**entry))
        self.deck = tuple(card_list)

    def calculate_total_price(self):
        self.evaluate_deck_price()
        running_total = 0.0
        for card in self.deck:
            running_total += card.price * card.quantity

        return running_total

    def evaluate_deck_price(self):
        if self.deck is None:
            self.populate_deck()
        _deck = self.deck
        for card in _deck:
            if card.price is None or card.price == 0 or (card.dirty is not None and card.dirty.lower() != 'n'):
                self.get_price(card)

        return _deck

    def get_price(self, card: Card):
        product_id = self.get_product_id(card)

        listing_request = util.create_request("listings")
        listing_request['filters']['term']['condition'] = [self.CARD_CONDITION_MAP[card.condition]]
        listing_request['filters']['term']['printing'] = [self.CARD_EDITION_MAP[card.edition]]
        listing_response = self.rest.listing(product_id, listing_request)

        if listing_response.ok:
            listing_response = listing_response.json()['results'][0]['results']
            for listing in listing_response:
                if not listing['directSeller']:
                    if listing['sellerShippingPrice'] != 0:
                        card.price = listing['price'] + listing['sellerShippingPrice']
                    else:
                        card.price = listing['price']
                    break

        return card.price

    def get_product_id(self, card: Card):
        fail_safe = 0
        check_next_page = True
        search_request = util.create_request("search")
        if card.rarity is not None:
            search_request["filters"]["term"]["rarityName"] = self.CARD_RARITY_MAP[card.rarity]
        size = int(search_request["size"])
        page = 0
        count = 0

        while check_next_page:
            check_next_page = False
            fail_safe += 1
            total_size = page * size
            search_request["from"] = total_size
            query = card.name.lower().replace(" ", "+")

            search_response = self.rest.search(query, search_request)

            results: dict
            if search_response.ok:
                results = search_response.json()['results']

                if len(results) > 0:
                    results = results[0]

                    if count == 0:
                        product_lines = results['aggregations']['productLineName']
                        for pl in product_lines:
                            if pl['urlValue'] == search_request['filters']['term']['productLineName'][0]:
                                count = pl['count']

                    sub_results = results['results']
                    for result in sub_results:
                        if result['customAttributes']['number'] == card.set_number:
                            return str(int(result['productId']))
                    if count - total_size > 0:
                        check_next_page = True

            if fail_safe > 4:
                break

        print(fail_safe)
        print(page)
        return None

    def write_to_excel(self, file_name: str = None):
        file = self.file_name if file_name is None else file_name
        wb = Workbook()
        wb = util.write_to_workbook(wb, self.deck)
        wb.save(self.file_path + file)

    def sanitize(self, entry: dict):
        print(entry)
        quantity = entry["quantity"]
        if quantity is None or quantity < 1 or not isinstance(quantity, int):
            entry["quantity"] = 1

        name = entry["name"].strip()
        name = re.sub(r'[^A-Za-z0-9 ]+', '', name)
        name = re.sub(r' +', ' ', name)
        entry["name"] = name

        condition = entry["condition"]
        if condition is None or condition not in self.CARD_CONDITION_MAP.keys():
            entry["condition"] = "LP"

        edition = entry["edition"]
        if edition is None or edition not in self.CARD_EDITION_MAP.keys():
            entry["edition"] = "U"

        rarity = entry["rarity"]
        if rarity is not None and rarity not in self.CARD_RARITY_MAP:
            entry["rarity"] = None
