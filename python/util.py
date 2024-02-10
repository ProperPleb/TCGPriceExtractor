import json
import config
from openpyxl import Workbook

VALID_REQUEST_TYPE = ["search", "listings"]


def create_request(request_type: str, file_path: str = None):
    req: json
    request_type = request_type.lower()
    if request_type in VALID_REQUEST_TYPE:
        file_name = request_type + ".json"
        path = file_path if file_path is not None else config.DEFAULT_PATH
        with open(path + file_name, 'r') as file:
            req = json.load(file)
    return req


def fill(sentence: str, a: chr, b: chr):
    if " " in sentence:
        sentence = sentence.replace(a, b)
    return sentence


def excel_mapper(wb: Workbook, column_names: list = None, table_offset_row: int = None, table_offset_col: int = None):
    dict_list = []
    fields = []
    table_width = None
    if column_names is not None and len(column_names) != 0:
        table_width = len(column_names)
        for column in column_names:
            fields.append(fill(column.strip().lower(), " ", "_"))

    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    max_col = sheet.max_column

    if table_offset_row is not None and table_offset_row >= 0 and table_offset_col is not None and table_offset_col >= 0:
        offset = {"row": table_offset_row, "col": table_offset_col}
    else:
        roc = {"row": None, "col": None}
        if table_offset_row is not None and table_offset_row >= 0:
            roc["row"] = table_offset_row
        elif table_offset_col is not None and table_offset_col >= 0:
            roc["col"] = table_offset_col
        offset = calculate_offset(sheet, max_row, max_col, table_width, roc)
    first_cell = sheet['A1'].offset(offset["row"], offset["col"])
    for i in range(max_row - offset["row"]):
        current_cell = first_cell.offset(i, 0)
        row = {}
        for j in range(max_col - offset["col"]):
            if i == 0 and column_names is None:
                fields.append(fill(str(current_cell.value).strip().lower(), " ", "_"))
            else:
                row.update({fields[j]: current_cell.value})
            current_cell = current_cell.offset(0, 1)
        if i != 0:
            dict_list.append(row)

    return dict_list


def calculate_offset(sheet, max_row: int = None, max_col: int = None, table_width: int = None, offset: dict = None):
    if max_row is None:
        max_row = sheet.max_row
    if max_col is None:
        max_col = sheet.max_column

    cell = sheet['A1']
    col_offset = offset["row"] if offset is not None else None
    row_offset = offset["col"] if offset is not None else None
    is_empty = True

    if col_offset is None:
        col_offset = 0
        # Check each cell from left to right until a populated cell is reached
        for i in range(max_col):
            if cell.value is None:
                col_offset += 1
                cell = cell.offset(0, 1)
            else:
                row_offset = 0
                is_empty = False
                break
    else:
        cell = cell.offset(0, col_offset)
        if cell.value is not None:
            is_empty = False

    # If a populated cell was not reached above, check each cell from top to bottom until a populated cell is reached
    if is_empty:
        cell = cell.offset(0, -1)
        if row_offset is None:
            row_offset = 0
            for j in range(max_row):
                if cell.value is None:
                    row_offset += 1
                    cell = cell.offset(1, 0)
                else:
                    is_empty = False
                    break
        else:
            if cell.offset(row_offset, 0).value is not None:
                is_empty = False

        if is_empty:
            raise Exception("Sheet is empty or entries out of bounds")

        # Recalculate col offset by backtracking until empty cell is found if table width not provided
        if table_width is None or table_width == 0:
            while cell.value is not None:
                cell = cell.offset(0, -1)
                col_offset -= 1
        else:
            col_offset = max_col - table_width

    return {"row": row_offset,
            "col": col_offset}


def write_to_workbook(wb: Workbook, tup: tuple):
    if len(wb.get_sheet_names()) > 0:
        wb.remove_sheet(wb.get_sheet_by_name(wb.get_sheet_names()[0]))
    wb.create_sheet('Sheet1', 0)
    sheet = wb.worksheets[0]
    first_cell = sheet['A1']
    current_cell = first_cell

    obj_type = type(tup[0])
    table_headers = vars(tup[0]).keys()
    for key in table_headers:
        current_cell.value = key
        current_cell = current_cell.offset(0, 1)
    index = 1
    for obj in tup:
        current_cell = first_cell.offset(index, 0)
        if type(obj) != obj_type:
            raise Exception("All objects in tuple must be of same type")
        for col in table_headers:
            current_cell.value = getattr(obj, col)
            current_cell = current_cell.offset(0, 1)
        index += 1
    return wb
